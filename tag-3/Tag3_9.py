import openpyxl as xls
workbook = xls.load_workbook(filename="mappe1.xlsx")
print(workbook.sheetnames)
workbook.create_sheet("Tabelle2")
print(workbook.sheetnames)
workbook.save(filename="mappe2.xls")