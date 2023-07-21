import openpyxl as xls
import warnings
from typing import List, Dict, Any

def get_excel_data(fname: str, sheet_name: str) -> List[Dict[str, Any]]:
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    wb = xls.open(fname)
    sheet = wb[sheet_name]

    header = [cell.value for cell in sheet[1]]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = [ cell for cell in row ]
        row_dict = dict(zip(header, values))
        data.append(row_dict)
    return data

def set_excel_data(fname: str, sheet_name: str, data: List[Dict[str, Any]]):
    """
    Schreibt Daten aus einer Liste von Dict in ein neues Excel Workbook

    :params:
        fname: str - der Dateiname
        sheet_name: str - der Sheetname
        data: List von Dict - die Daten für Excel
    :returns: None
    """
    wb = xls.Workbook()
    sheet = wb.active
    sheet.title = sheet_name  # Create a new sheet with the provided name

    if not data:  # if data is empty
        wb.save(fname)  # Save the workbook

    # Write headers (keys of the first dict)
    headers = list(data[0].keys())

    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    # Write data rows
    for row_num, row_data in enumerate(data, start=2):
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=row_num, column=col_num, value=row_data.get(header))

    wb.save(fname)  # Save the workbook


if __name__ == "__main__": # Spielwiese, aber kein Test!!!
    import pprint
    sheet_name = "Titanic_ohne_Name"
    fname = r"E:\Workspaces\Kurse\aktueller-kurs\Material\Titanic_splitted_names.xlsx"

    data = get_excel_data(fname, sheet_name)
    pprint.pprint(data[:2])
    set_excel_data("Neue_Daten.xlsx", "Titanic", data)

    import data_io_tools as io
    io.write_json_data("Titanic", data)
