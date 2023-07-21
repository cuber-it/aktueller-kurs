#-- Excel Section
import openpyxl as xls
import warnings
from typing import List, Dict, Any

def read_excel_data(fname: str, sheet_name: str) -> List[Dict[str, Any]]:
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

    wb = xls.open(fname)
    sheet = wb[sheet_name]

    header = [cell.value for cell in sheet[1]]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = [ cell for cell in row ]
        row_dict = dict(zip(header, values))
        data.append(row_dict)
    return data

def write_excel_data(fname: str, sheet_name: str, data: List[Dict[str, Any]]):
    """
    Schreibt Daten aus einer Liste von Dict in ein neues Excel Workbook

    :params:
        fname: str - der Dateiname
        sheet_name: str - der Sheetname
        data: List von Dict - die Daten für Excel
    :returns: None
    """
    wb = xls.Workbook()
    sheet = wb.active
    sheet.title = sheet_name  # Create a new sheet with the provided name

    if data:  # if data is empty
        # Write headers (keys of the first dict)
        headers = list(data[0].keys())

        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_num, value=header)

        # Write data rows
        for row_num, row_data in enumerate(data, start=2):
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=row_num, column=col_num, value=row_data.get(header))
    wb.save(fname)  # Save the workbook

#-- SqlLite-Section
import sqlite3
import json

def read_sqlite_data(database, table_name):
    connection = sqlite3.connect(database)
    cursor = connection.cursor()
    table = cursor.execute(f"SELECT * FROM {table_name}")

    # List Comprehension: names = [description[0] for description in cursor.description]
    names = []
    for description in cursor.description:
        names.append(description[0])

    # List Comprehension: daten = [dict(zip(names, row)) for row in table.fetchall()]
    daten = []
    for row in table.fetchall():
        daten.append(dict(zip(names, row,)))
    return daten


def write_sqlite_data(database, table_name, data):
    def type_map(variable):
        return {
            "int": 'INTEGER',
            "float": 'REAL',
            "str": 'TEXT',
            "bytes": 'BLOB',
            "bool": 'INTEGER',  # SQLite uses integer for boolean values, 0 (false) and 1 (true)
        }.get(type(variable).__name__, "TEXT")

    connection = sqlite3.connect(database)
    cursor = connection.cursor()

    #
    cmd = f"CREATE TABLE {table_name} ("
    #
    cmd += ", ".join([f"{key} {type_map(data[0][key])}" for key in data[0].keys()])
    cmd += ")"
    #print("DEBUG: ", cmd)
    cursor.execute(cmd)

    cmd = f"INSERT INTO {table_name} ("
    cmd += ", ".join(list(data[0].keys()))
    cmd += ") VALUES ("
    #
    cmd += ", ".join(["?"] * len(list(data[0].keys())))
    cmd +=")"
    #print("DEBUG: ", cmd)

    for row in data:
        cursor.execute(cmd, list(row.values()))
    connection.commit()

#-- JSON-Section
def read_json_data(fname):
    pass

def write_json_data(fname, daten, indent=4):
    fname += ".json"
    with open(fname, "w") as fp:
        json.dump(daten, fp, indent=indent)
